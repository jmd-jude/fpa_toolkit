"""
Box Manifest → Formatted Case Report (Excel)

Reads box_manifest.csv and generates a formatted Excel report
organized by plaintiff / top-level folder.

Usage:
    python report.py

Output:
    box_case_report.xlsx

Configuration (optional — edit below):
    CASE_NAME      Override the case name in the report header.
                   Defaults to the root folder name from the manifest.
    DATE_RECEIVED  Date to show in the header (free text).
    SKIP_FOLDERS   Top-level folder names to exclude from the report.
"""

import csv
import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# --- CONFIGURE (optional) ---
CASE_NAME = ''           # Leave blank to use root folder name from manifest
DATE_RECEIVED = ''       # e.g. '03/17/2026' — leave blank to omit
SKIP_FOLDERS = [         # Top-level folders to exclude from report
    '**Records Received in Original Format (by Attorney)**',
]
INPUT_FILE = ''   # Leave blank to auto-detect the most recent *_manifest.csv
OUTPUT_FILE = ''  # Leave blank to auto-name from the manifest filename
# ----------------------------

# ── Colours (matching Apara report style) ───────────────────────────────────
GREEN   = PatternFill('solid', fgColor='4CAF50')   # section header
SUBHEAD = PatternFill('solid', fgColor='C8E6C9')   # subsection header (light green)
YELLOW  = PatternFill('solid', fgColor='FFEB3B')   # case summary header
BLACK   = PatternFill('solid', fgColor='212121')   # column header bar
PINK    = PatternFill('solid', fgColor='F8BBD0')   # subtotal row
LTGRAY  = PatternFill('solid', fgColor='F5F5F5')   # alternating row

WHITE_BOLD  = Font(bold=True, color='FFFFFF')
BLACK_BOLD  = Font(bold=True, color='000000')
YELLOW_BOLD = Font(bold=True, color='212121')


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def parse_date_from_filename(name):
    """Extract YYYY-MM-DD date prefix from filenames like '2024-03-26 Description.pdf'"""
    stem = os.path.splitext(name)[0]
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', stem)
    if m:
        # Convert to M/D/YYYY display format
        y, mo, d = m.group(1).split('-')
        return f'{int(mo)}/{int(d)}/{y}'
    return ''


def top_level_section(path, case_root):
    """
    Extract the top-level folder (section) and optional subsection from a path.
    e.g.  'Case/01_Plaintiff/subdir/file.pdf'  →  ('01_Plaintiff', 'subdir')
    """
    parts = path.split('/')
    # parts[0] = case root, parts[1] = section, parts[2..] = subsection + file
    if len(parts) < 3:
        return case_root, ''
    section = parts[1]
    subsection = '/'.join(parts[2:-1]) if len(parts) > 3 else ''
    return section, subsection


def load_manifest(filepath):
    rows = []
    with open(filepath, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows


def group_by_section(rows, skip_folders):
    """
    Returns:
        case_root  — name of the top-level case folder
        sections   — dict: section_name → list of file rows
                     each row has an added '_subsection' key
    """
    case_root = ''
    sections = defaultdict(list)

    for row in rows:
        path = row['Path']
        parts = path.split('/')
        if not case_root and len(parts) > 0:
            case_root = parts[0]

        section, subsection = top_level_section(path, case_root)
        if any(section.startswith(s) for s in skip_folders):
            continue
        row_copy = dict(row)
        row_copy['_subsection'] = subsection
        sections[section].append(row_copy)

    return case_root, sections


def write_duplicates_sheet(wb, rows):
    """Add a 'Duplicates' worksheet listing all files that share Name + Size."""
    dup_rows = [r for r in rows if r.get('Duplicate', 'No') == 'Yes']
    if not dup_rows:
        return  # nothing to add

    ws = wb.create_sheet(title='Duplicates')

    # Column widths
    ws.column_dimensions['A'].width = 45   # File Name
    ws.column_dimensions['B'].width = 10   # Type
    ws.column_dimensions['C'].width = 12   # Size
    ws.column_dimensions['D'].width = 65   # Folder Path

    # Header row
    headers = ['FILE NAME', 'TYPE', 'SIZE', 'FOLDER PATH']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = BLACK
        c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
    ws.row_dimensions[1].height = 18

    # Summary note
    unique_pairs = len(set((r['Name'], r['Size (KB)']) for r in dup_rows))
    ws.merge_cells('A2:D2')
    note = ws['A2']
    note.value = (f'{len(dup_rows)} duplicate instances found '
                  f'({unique_pairs} unique file names/sizes appearing in multiple folders)')
    note.fill = PINK
    note.font = BLACK_BOLD
    note.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[2].height = 16

    sorted_dups = sorted(dup_rows, key=lambda r: (r['Name'], r['Size (KB)']))
    for i, row in enumerate(sorted_dups, 1):
        r_num = i + 2
        ext = row.get('Extension', '').lstrip('.').upper() or '—'
        size = row.get('Size (KB)', '')
        if size and size != 'N/A':
            kb = float(size)
            size = f'{kb/1024:.1f} MB' if kb >= 1024 else f'{kb:.0f} KB'

        row_fill = LTGRAY if i % 2 == 1 else PatternFill()
        vals = [row['Name'], ext, size, row.get('Folder', '')]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r_num, column=col, value=val)
            c.fill = row_fill
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal='left' if col in (1, 4) else 'center',
                vertical='center',
                wrap_text=(col in (1, 4)),
            )
        ws.row_dimensions[r_num].height = 15


def write_report(case_root, sections, rows, output_file=None, case_name_override=None, date_received_override=None):
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = 'Case Report'

    # Column widths — file name column sized to data, capped at 120
    max_name_len = max((len(r['Name']) for r in rows), default=40)
    name_col_width = min(max(max_name_len + 4, 30), 120)

    ws.column_dimensions['A'].width = 6             # #
    ws.column_dimensions['B'].width = name_col_width # File Name
    ws.column_dimensions['C'].width = 10            # File Type
    ws.column_dimensions['D'].width = 14            # Total Pages
    ws.column_dimensions['E'].width = 16            # Document Date
    ws.column_dimensions['F'].width = 12            # Size
    ws.column_dimensions['G'].width = 30            # Notes

    case_name = case_name_override or CASE_NAME or case_root.upper()
    date_recv  = date_received_override if date_received_override is not None else DATE_RECEIVED
    total_files = sum(1 for r in rows if r['Extension'] not in ('', 'no extension'))
    total_pages_known = sum(
        int(r['Page Count']) for r in rows if r['Page Count'] not in ('N/A', '')
    )
    pages_display = f"{total_pages_known:,}+ (partial — see N/A rows)"

    # ── Case header block ────────────────────────────────────────────────────
    def merge_write(row_num, value, fill, font, alignment='center'):
        ws.merge_cells(f'A{row_num}:G{row_num}')
        cell = ws[f'A{row_num}']
        cell.value = value
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal=alignment, vertical='center')
        ws.row_dimensions[row_num].height = 20

    merge_write(1, case_name,                    GREEN,  Font(bold=True, size=14, color='FFFFFF'))
    if date_recv:
        merge_write(2, f'Date Received: {date_recv}', YELLOW, Font(bold=True, size=11))
    merge_write(3, f'TOTAL FILES: {total_files:,}',       BLACK,  WHITE_BOLD)
    merge_write(4, f'KNOWN PAGES: {total_pages_known:,} (non-PDF files show N/A)', BLACK, WHITE_BOLD)

    current_row = 6

    def write_col_headers(r):
        headers = ['#', 'FILE NAME', 'TYPE', 'TOTAL PAGES', 'DOCUMENT DATE', 'SIZE', 'NOTES']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=col, value=h)
            c.fill = BLACK
            c.font = WHITE_BOLD
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()
        ws.row_dimensions[r].height = 18

    # ── Sections ─────────────────────────────────────────────────────────────
    for section_name, section_rows in sorted(sections.items()):
        # Find folder URL from the first file sitting directly in this section
        section_folder_url = next(
            (r.get('Folder URL', '') for r in section_rows if r.get('_subsection', '') == ''),
            ''
        )

        # Section header
        ws.merge_cells(f'A{current_row}:G{current_row}')
        c = ws[f'A{current_row}']
        c.value = section_name
        c.fill = GREEN
        if section_folder_url:
            c.hyperlink = section_folder_url
            c.font = Font(bold=True, color='FFFFFF', size=11, underline='single')
        else:
            c.font = Font(bold=True, color='FFFFFF', size=11)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        write_col_headers(current_row)
        current_row += 1

        section_pages = 0
        section_na = 0
        item_num = 1

        # Group rows by subsection, preserving order of first appearance
        sub_order = []
        sub_groups = defaultdict(list)
        for row in section_rows:
            sub = row.get('_subsection', '')
            if sub not in sub_groups:
                sub_order.append(sub)
            sub_groups[sub].append(row)

        has_subsections = len(sub_order) > 1

        for sub_name in sub_order:
            sub_rows = sub_groups[sub_name]

            # Subsection header — only shown when there are multiple subsections
            if has_subsections:
                display = sub_name if sub_name else '(root)'
                sub_folder_url = sub_rows[0].get('Folder URL', '') if sub_rows else ''
                ws.merge_cells(f'A{current_row}:G{current_row}')
                c = ws[f'A{current_row}']
                c.value = f'  {display}'
                c.fill = SUBHEAD
                if sub_folder_url:
                    c.hyperlink = sub_folder_url
                    c.font = Font(bold=True, color='1B5E20', size=10, underline='single')
                else:
                    c.font = Font(bold=True, color='1B5E20', size=10)
                c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                ws.row_dimensions[current_row].height = 16
                current_row += 1

            for i, row in enumerate(sub_rows):
                page_count = row['Page Count']
                if page_count not in ('N/A', ''):
                    section_pages += int(page_count)
                else:
                    section_na += 1

                doc_date = parse_date_from_filename(row['Name']) or row.get('Modified', '')
                size = row.get('Size (KB)', '')
                if size and size != 'N/A':
                    kb = float(size)
                    size = f'{kb/1024:.1f} MB' if kb >= 1024 else f'{kb:.0f} KB'

                file_type = row.get('Extension', '').lstrip('.').upper() or '—'
                row_fill = LTGRAY if item_num % 2 == 1 else PatternFill()
                vals = [item_num, row['Name'], file_type, page_count, doc_date, size, '']

                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.fill = row_fill
                    c.border = thin_border()
                    c.alignment = Alignment(
                        horizontal='center' if col != 2 else 'left',
                        vertical='center',
                        wrap_text=(col == 2)
                    )
                    if col == 2:
                        file_url = row.get('File URL', '')
                        if file_url:
                            c.hyperlink = file_url
                            c.font = Font(color='1565C0', underline='single')
                lines = max(1, -(-len(row['Name']) // name_col_width))  # ceiling div
                ws.row_dimensions[current_row].height = max(15, lines * 14)
                item_num += 1
                current_row += 1

        # Section subtotal row
        ws.merge_cells(f'A{current_row}:B{current_row}')
        subtotal_label = ws[f'A{current_row}']
        subtotal_label.value = f'Subtotal — {len(section_rows)} files'
        subtotal_label.fill = PINK
        subtotal_label.font = BLACK_BOLD
        subtotal_label.alignment = Alignment(horizontal='right', vertical='center')

        pages_cell = ws.cell(row=current_row, column=3)
        na_note = f' (+{section_na} N/A)' if section_na else ''
        pages_cell.value = f'{section_pages:,}{na_note}'
        pages_cell.fill = PINK
        pages_cell.font = BLACK_BOLD
        pages_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(4, 8):
            ws.cell(row=current_row, column=col).fill = PINK

        ws.row_dimensions[current_row].height = 16
        current_row += 2   # blank row between sections

    write_duplicates_sheet(wb, rows)

    if output_file is None:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    else:
        wb.save(output_file)
        print(f'\nReport saved → {output_file}')
        print(f'  Sections:     {len(sections)}')
        print(f'  Total files:  {total_files:,}')
        print(f'  Known pages:  {total_pages_known:,}')


def resolve_input_file():
    """Auto-detect the most recently modified *_manifest.csv if INPUT_FILE not set."""
    if INPUT_FILE:
        return INPUT_FILE
    candidates = [f for f in os.listdir('.') if f.endswith('_manifest.csv')]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def main():
    input_file = resolve_input_file()
    if not input_file or not os.path.exists(input_file):
        print('ERROR: No manifest CSV found. Run manifest.py first.')
        return

    slug = input_file.replace('_manifest.csv', '')
    output_file = OUTPUT_FILE or f'{slug}_report.xlsx'

    print(f'Reading: {input_file}')
    rows = load_manifest(input_file)
    case_root, sections = group_by_section(rows, SKIP_FOLDERS)
    write_report(case_root, sections, rows, output_file)


if __name__ == '__main__':
    main()
